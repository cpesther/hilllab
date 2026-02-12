# Christopher Esther, Hill Lab, 2/6/2026
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

from ..migration.plate_extract_result import plate_extract_result
from ..utilities.remove_outliers import remove_outliers

def plate_generate_report(bundle, groups):

    """
    Generates and saves a summary of 1D results to an Excel file in the
    same folder as the raw data. 

    ARGUMENTS:
        bundle (migration.Bundle): the data bundle after processing
        groups (dict): a dictionary of which columns should be grouped
            together, where the keys are the names of the groups as strings
            and the value is a list of the column numbers included in the
            group.
    """

    # Determine the maximum column length for spacing
    max_n = 0
    for key in groups.keys():
        max_n = max(max_n, len(groups[key]))

    # Generate raw data tables for summary stats
    index = [f'n{i+1}' for i in range(max_n)]
    eta_table = pd.DataFrame({'capillary': index}).set_index('capillary')
    D_table = pd.DataFrame({'capillary': index}).set_index('capillary')

    # Iterate over each key in the group names
    for key in groups.keys():

        # Some lists to save the values in this column
        eta_values = []
        D_values = []
        
        # Iterate over every column in this group
        for column in groups[key]:
            
            eta_values.append(plate_extract_result(bundle, column, 'eta', 'mean', end_only=True, end_hours=24))
            D_values.append(plate_extract_result(bundle, column, 'D', 'mean', end_only=True, end_hours=24))

        # Append enough nan values to match length
        if len(eta_values) < max_n:
            eta_values.extend(np.repeat(None, (max_n - len(eta_values))))
            D_values.extend(np.repeat(None, (max_n - len(D_values))))
            
        # Save these values to their respective tables
        eta_table[key] = eta_values
        D_table[key] = D_values

    # Create tables for data with outliers removed
    eta_table_clean = pd.DataFrame({'capillary': index}).set_index('capillary')
    D_table_clean = pd.DataFrame({'capillary': index}).set_index('capillary')

    for column in eta_table.columns:

        # Pull values from table and remove outliers
        eta_values = eta_table[column].to_numpy(dtype=object)
        D_values = D_table[column].to_numpy(dtype=object)
        _, _, outlier_indices = remove_outliers(array=eta_values, info=False, paste_mode=False)

        # Replace outliers with string
        for i in outlier_indices:
            eta_values[i] = 'OUTLIER'
            D_values[i] = 'OUTLIER'

        eta_table_clean[column] = eta_values  # save to clean table
        D_table_clean[column] = D_values

    # Determine report name and path
    report_name = f'REPORT_{Path(bundle.data.path).stem}.xlsx'
    report_path = os.path.join(Path(bundle.data.path).parent, report_name)

    # Calculate these early for placement of tables in pandas writer
    mean_row = eta_table.shape[0] + 3
    median_row = eta_table.shape[0] + 4
    stdev_row = eta_table.shape[0] + 5
    raw_data_row = stdev_row + 2

    # Output data to Excel file
    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        eta_table_clean.to_excel(
            writer,
            startrow=1,
            sheet_name='Report',
        )

        D_table_clean.to_excel(
            writer,
            startrow=1,
            sheet_name='Report',
            startcol=len(eta_table.columns) + 2
        )
        eta_table.to_excel(
            writer,
            startrow=raw_data_row,
            sheet_name='Report',
        )

        D_table.to_excel(
            writer,
            startrow=raw_data_row,
            sheet_name='Report',
            startcol=len(eta_table.columns) + 2
        )

    # A helper function for convering column numbers to Excel letters
    def colnum_to_letter(n):
        """Convert 1-based Excel column number to letter"""
        result = ''
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result
    
    # Calculate start and end columns for each table
    total_eta_cols = eta_table.shape[1] + 1  # +1 for index
    eta_start_col = 'A'
    eta_end_col = colnum_to_letter(total_eta_cols)  # last used column

    # Leave blank column between tables (+1 for blank, +1 for index)
    D_start_col = colnum_to_letter(total_eta_cols + 2)
    D_end_col = colnum_to_letter(total_eta_cols + 2 + D_table.shape[1] + 1 - 1)

    # Load workbook for some additional formatting
    wb = load_workbook(report_path)
    ws = wb['Report']

    # First, scan the clean data tables and highlight any outlier values
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)

    # Loop over range of cells (for example B2:E10)
    for row in ws['A1':f'{D_end_col}{mean_row}']:
        for cell in row:
            if cell.value == 'OUTLIER':
                cell.fill = red_fill
                cell.font = white_font

    # Formatting titles for normalized data tables
    ws.merge_cells(f'{eta_start_col}1:{eta_end_col}1')
    ws.merge_cells(f'{D_start_col}1:{D_end_col}1')
    ws['A1'] = 'Eta (mPa•s)'
    ws[f'{D_start_col}1'] = 'Diffusion Coefficient (m²s)'

    ws['A1'].font = Font(color='4B54CC', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{D_start_col}1'].font = Font(color='4B54CC', bold=True, size=14)
    ws[f'{D_start_col}1'].alignment = Alignment(horizontal='center', vertical='center')

    # Formatting titles for the raw data tables
    raw_data_title_row = stdev_row + 2
    ws.merge_cells(f'{eta_start_col}{raw_data_title_row}:{eta_end_col}{raw_data_title_row}')
    ws.merge_cells(f'{D_start_col}{raw_data_title_row}:{D_end_col}{raw_data_title_row}')
    ws[f'{eta_start_col}{raw_data_title_row}'] = 'Eta (mPa•s) (raw)'
    ws[f'{D_start_col}{raw_data_title_row}'] = 'Diffusion Coefficient (m²s) (raw)'

    ws[f'{eta_start_col}{raw_data_title_row}'].font = Font(color='FF0000', bold=True, size=14)
    ws[f'{eta_start_col}{raw_data_title_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{D_start_col}{raw_data_title_row}'].font = Font(color='FF0000', bold=True, size=14)
    ws[f'{D_start_col}{raw_data_title_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # Generate the blocks of summary formulas
    right_align = Alignment(horizontal='right', vertical='center')

    def excel_row_cells(start_col_num, num_cols, row_num):
        """Return list of Excel cell addresses for a row."""
        return [f'{colnum_to_letter(start_col_num + i)}{row_num}' for i in range(num_cols)]

    def write_metric_formulas(ws, start_col_num, num_cols, data_start_row, mean_row, median_row, stdev_row):
        """
        Write formulas for Mean, Median, StDev for a table starting at start_col_num with num_cols columns.
        data_start_row: first row of actual data (below header)
        """
        center_align = Alignment(horizontal='center', vertical='center')

        # Add in the labels
        labels = ['Mean', 'Median', 'StDev']
        for i, label in enumerate(labels):
            label_row = mean_row + i
            cell = ws[f'{colnum_to_letter(start_col_num - 1)}{label_row}']
            cell.value = label
            cell.alignment = right_align
            cell.font = Font(bold=True, size=11)
        
        # Loop through each column
        for i in range(num_cols):
            col_letter = colnum_to_letter(start_col_num + i)
            data_range = f'{col_letter}{data_start_row}:{col_letter}{data_start_row + num_rows - 1}'

            ws[f'{col_letter}{mean_row}']   = f'=AVERAGE({data_range})'
            ws[f'{col_letter}{median_row}'] = f'=MEDIAN({data_range})'
            ws[f'{col_letter}{stdev_row}']  = f'=STDEV({data_range})'
            # Apply center alignment
            for row in [mean_row, median_row, stdev_row]:
                ws[f'{col_letter}{row}'].alignment = center_align

            ws[f'{col_letter}{mean_row}'].font = Font(bold=True, size=11)

    num_rows = len(eta_table)   # number of data rows in each table
    data_start_row = 3          # first row with actual data, assuming row 1 is header

    # Eta table
    eta_start_col_num = 2
    total_eta_cols = eta_table.shape[1]
    write_metric_formulas(ws, eta_start_col_num, total_eta_cols, data_start_row, mean_row, median_row, stdev_row)

    # D table
    D_start_col_num = total_eta_cols + 4
    total_D_cols = D_table.shape[1]
    write_metric_formulas(ws, D_start_col_num, total_D_cols, data_start_row, mean_row, median_row, stdev_row)

    # Raw eta and D tables
    raw_offset = data_start_row + eta_table.shape[0] + 3
    raw_data_start_row = data_start_row + raw_offset
    write_metric_formulas(ws, eta_start_col_num, total_eta_cols, raw_data_start_row, 
                        mean_row + raw_offset, 
                        median_row + raw_offset, 
                        stdev_row+ raw_offset)

    write_metric_formulas(ws, D_start_col_num, total_D_cols, raw_data_start_row, 
                        mean_row + raw_offset, 
                        median_row + raw_offset, 
                        stdev_row + raw_offset)
    
    # Merge cells along bottom for disclaimer message
    disclaimer_cells = excel_row_cells(1, D_start_col_num + total_D_cols - 1, stdev_row + 2 + raw_offset)
    ws.merge_cells(f'{disclaimer_cells[0]}:{disclaimer_cells[-1]}')
    ws[disclaimer_cells[0]] = 'Report is automatically generated.'
    ws[disclaimer_cells[0]].font = Font(italic=True, size=10)
    ws[disclaimer_cells[0]].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(report_path)

    print(f'Report generated at {report_path}')
